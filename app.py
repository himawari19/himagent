import os
import re
import json
import traceback
import hashlib
import threading
import sqlite3
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from datetime import datetime
from flask import Flask, render_template, request, send_file, jsonify
from werkzeug.utils import secure_filename
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import google.generativeai as genai
from pydantic import BaseModel, Field
from typing import Literal
from cryptography.fernet import Fernet
from uuid import uuid4
from core.cache_store import get_generation_cache, set_generation_cache
from core.file_ops import (
    delete_generated_file as core_delete_generated_file,
    find_file_in_outputs as core_find_file_in_outputs,
    list_generated_tree,
    safe_filename,
)
from core.quality import evaluate_quality
from core.upload_validation import validate_screenshot_payload
from core.prompts import get_system_prompt, get_few_shot_hint
from core.jobs import (
    _GENERATION_LOCK,
    _GENERATION_JOBS,
    _GENERATION_CACHE,
    _GENERATION_ACTIVE_BY_CACHE,
    _GENERATION_DB_PATH,
    _set_job_state,
    _get_job_state,
)
from core.providers import (
    google_exceptions,
    detect_provider,
    pil_to_base64_jpeg,
    parse_api_error,
    parse_gemini_exception,
    call_openai_compatible_api,
    call_claude_api,
)
from core.router import (
    ROUTER_BASE_URL,
    ROUTER_DEFAULT_MODEL,
    normalize_router_model,
    format_router_ping_error,
    ping_router_model_once,
    test_router_model_route,
    fetch_router_models,
)
from core.stats import (
    get_outputs_dir as _outputs_dir,
    build_outputs_file_map as _build_outputs_file_map,
    filename_to_module_label as _filename_to_module_label,
    count_test_cases_from_workbook as _count_test_cases_from_workbook,
)
from core.exports import export_workbook
from core.preview import get_file_preview
from core.detection import detect_module_from_screenshot, resolve_api_key

# -----------------------------------------------------------------------------
# EPHEMERAL ENCRYPTION KEY (generated fresh each server start)
# -----------------------------------------------------------------------------
_FERNET_KEY = Fernet.generate_key()
_fernet = Fernet(_FERNET_KEY)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'uploads')
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # Disable static file caching in development
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB max upload
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# -----------------------------------------------------------------------------
# OUTPUT DIRECTORY INIT
# -----------------------------------------------------------------------------
def init_outputs_dir():
    root_dir = os.path.abspath(os.path.dirname(__file__))
    outputs_dir = os.path.join(root_dir, 'outputs')
    os.makedirs(outputs_dir, exist_ok=True)

init_outputs_dir()


def _normalize_ai_response(data):
    """Coerce AI responses into the dict shape expected by the generators."""
    if isinstance(data, list):
        return {'test_cases': data}
    if not isinstance(data, dict):
        raise Exception(f"Unexpected AI response type: {type(data).__name__}")

    if 'test_cases' in data:
        return data

    for value in data.values():
        if isinstance(value, dict) and 'test_cases' in value:
            return value

    return data


# -----------------------------------------------------------------------------
# PYDANTIC SCHEMAS FOR STRUCTURED GEMINI OUTPUT
# -----------------------------------------------------------------------------

class TestCaseSchema(BaseModel):
    tc_id: str = Field(description="Unique sequential identifier for the test case (e.g. TC-L-001, TC-L-002, etc.)")
    scenario: str = Field(description="Descriptive scenario title. MUST follow: '[Feature Area Element] - [Test Description]' in Title Case. Example: 'Navbar Logo - Verify clicking Vania logo navigates to homepage'.")
    case_type: Literal["Positive", "Negative", "Boundary"] = Field(description="Type of test case: Positive (valid flow), Negative (error handling, invalid input, security), Boundary (limits, empty inputs, stress states)")
    precondition: str = Field(description="Prerequisite state before executing the steps")
    steps: str = Field(description="Numbered step-by-step actions (e.g. '1. Click X\n2. Enter Y')")
    expected: str = Field(description="Precise expected outcome")

class ElementSchema(BaseModel):
    area: str = Field(description="UI area / component section (e.g., Navbar, Header, Sidebar, Form Inputs, Advanced options)")
    element_name: str = Field(description="Name of the UI component")
    element_type: str = Field(description="Component type (e.g., Dropdown, Button, Text Input, Checkbox, Link)")
    interactions: int = Field(description="Assumed number of interactions/clicks required for testing this element")

class ChecklistSchema(BaseModel):
    category: str = Field(description="Checklist Category (UI/UX Layout, Functional, Accessibility (WCAG), Security, Cross-Browser, Mobile Responsive, Performance, Session/Auth)")
    checklist_item: str = Field(description="Manual checklist item description starting with 'Verify that...'")




def _make_generation_cache_key(payload):
    hasher = hashlib.sha256()
    key_parts = [
        payload.get("page_title", ""),
        payload.get("id_prefix", ""),
        payload.get("model_name", ""),
        payload.get("instructions", ""),
        payload.get("provider", ""),
        payload.get("gen_depth", ""),
        "1" if payload.get("generate_script") else "0",
    ]
    for context in payload.get("screen_contexts", []):
        key_parts.append(context)
    for part in key_parts:
        hasher.update(str(part).encode("utf-8", errors="ignore"))
        hasher.update(b"\0")
    for file_item in payload.get("files", []):
        hasher.update(file_item.get("filename", "").encode("utf-8", errors="ignore"))
        hasher.update(b"\0")
        hasher.update(file_item.get("bytes", b""))
        hasher.update(b"\0")
    return hasher.hexdigest()


def _run_generation_pipeline(payload, job_id=None):
    uploaded_files = payload["files"]
    page_title = payload["page_title"]
    id_prefix = payload["id_prefix"]
    model_name = payload["model_name"]
    instructions = payload["instructions"]
    screen_contexts = payload.get("screen_contexts", []) or []
    user_api_key = payload["api_key"]
    req_provider = payload["provider"]
    gen_depth = payload["gen_depth"]
    generate_script = payload["generate_script"]

    if job_id:
        _set_job_state(job_id, status="running", message="Preparing upload data...", progress=5)

    sys_prompt = get_system_prompt(gen_depth)
    provider = detect_provider(req_provider, model_name, user_api_key)
    model_name = normalize_router_model(provider, model_name)

    api_key = user_api_key if (user_api_key and user_api_key.lower() != 'env') else ""
    if not api_key:
        if provider == 'gemini':
            api_key = os.environ.get("GEMINI_API_KEY", "")
        elif provider == 'openai':
            api_key = os.environ.get("OPENAI_API_KEY", "")
        elif provider == 'claude':
            api_key = os.environ.get("ANTHROPIC_API_KEY", "") or os.environ.get("CLAUDE_API_KEY", "")
        elif provider == 'mimo':
            api_key = os.environ.get("MIMO_API_KEY", "") or os.environ.get("XIAOMI_API_KEY", "")
        elif provider == 'deepseek':
            api_key = os.environ.get("DEEPSEEK_API_KEY", "")
        elif provider == 'grok':
            api_key = os.environ.get("XAI_API_KEY", "") or os.environ.get("GROK_API_KEY", "")
        elif provider == 'mistral':
            api_key = os.environ.get("MISTRAL_API_KEY", "")
        elif provider == '9router':
            api_key = ""

    if not api_key and provider != '9router':
        raise Exception(f"{provider.upper()} API Key is required. Please set it in your environment variables or paste it in the form.")

    pil_images = []
    try:
        if job_id:
            _set_job_state(job_id, message="Reading and resizing screenshots...", progress=15)
        max_images = 3 if gen_depth == 'ultra' else (5 if gen_depth == 'fast' else 10)
        max_side  = 320 if gen_depth == 'ultra' else (512 if gen_depth == 'fast' else 768)
        files_to_process = uploaded_files[:max_images]

        def _load_image(file_item):
            img = Image.open(BytesIO(file_item.get("bytes", b"")))
            if img.mode in ("RGBA", "P"):
                img = img.convert("RGB")
            w, h = img.size
            if max(w, h) > max_side:
                scale = max_side / max(w, h)
                img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
            return img

        with ThreadPoolExecutor(max_workers=min(len(files_to_process), 4)) as ex:
            futures = {ex.submit(_load_image, f): i for i, f in enumerate(files_to_process)}
            ordered = [None] * len(futures)
            for fut in as_completed(futures):
                ordered[futures[fut]] = fut.result()
        pil_images = [img for img in ordered if img is not None]
    except Exception as e:
        raise Exception(f"Failed to process images: {str(e)}")

    if gen_depth == "ultra":
        prompt_prefix = (
            f"Analyze {len(pil_images)} sequential UI screenshots for '{page_title}'. "
            f"Return only test_cases. Use prefix '{id_prefix}'. "
        )
    elif generate_script:
        prompt_prefix = (
            f"Analyze {len(pil_images)} sequential UI screenshots for '{page_title}'. "
            f"Return test_cases, elements, and checklist when available. Use prefix '{id_prefix}'. "
        )
    else:
        prompt_prefix = (
            f"Analyze {len(pil_images)} sequential UI screenshots for '{page_title}'. "
            f"Return only test_cases; elements and checklist may be omitted. Use prefix '{id_prefix}'. "
        )

    context_lines = []
    for idx, context in enumerate(screen_contexts[:len(pil_images)], 1):
        context = str(context or '').strip()
        if context:
            context_lines.append(f"{idx}. Screenshot {idx}: {context}")
    sequence_context = ""
    if context_lines:
        sequence_context = (
            "Screenshot sequence context supplied by the user:\n"
            + "\n".join(context_lines)
            + "\nTreat these screenshots as one connected user journey when the context describes transitions or clicks. "
        )

    prompt = (
        f"{prompt_prefix}"
        f"The screenshots are ordered sequentially from first to last. "
        f"{sequence_context}"
        f"Additional context or constraints: {instructions if instructions else 'None'}"
        f" Generate all test plan content in English."
        f"{get_few_shot_hint()}"
    )

    data = None
    if job_id:
        _set_job_state(job_id, message="Running AI analysis...", progress=40)
    # Images already resized — encode to base64 JPEG once, reuse across all providers
    base64_images = [pil_to_base64_jpeg(img) for img in pil_images]

    if provider == 'gemini':
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel(model_name)
        response = model.generate_content(
            contents=[prompt, *pil_images, sys_prompt],
            generation_config=genai.GenerationConfig(response_mime_type="application/json"),
            request_options={"timeout": 300}
        )
        data = json.loads(response.text)
    elif provider == 'openai':
        data = call_openai_compatible_api("https://api.openai.com/v1", model_name, api_key, sys_prompt, prompt, base64_images, provider="OpenAI")
    elif provider == 'claude':
        data = call_claude_api(model_name, api_key, sys_prompt, prompt, base64_images)
    elif provider == 'mimo':
        data = call_openai_compatible_api("https://api.xiaomimimo.com/v1", model_name, api_key, sys_prompt, prompt, base64_images, provider="MiMo")
    elif provider == 'deepseek':
        data = call_openai_compatible_api("https://api.deepseek.com", model_name, api_key, sys_prompt, prompt, base64_images, provider="DeepSeek")
    elif provider == 'grok':
        data = call_openai_compatible_api("https://api.x.ai/v1", model_name, api_key, sys_prompt, prompt, base64_images, provider="Grok")
    elif provider == 'mistral':
        data = call_openai_compatible_api("https://api.mistral.ai/v1", model_name, api_key, sys_prompt, prompt, base64_images, provider="Mistral")
    elif provider == '9router':
        data = call_openai_compatible_api(ROUTER_BASE_URL, model_name, api_key, sys_prompt, prompt, base64_images, provider="9Router")

    data = _normalize_ai_response(data)
    if not data:
        raise Exception("No data returned from AI Model engine.")

    module_name = page_title.strip() or 'the module'

    repair_stats = {
        'filled_fields': 0,
        'case_type_fixed': 0,
        'steps_numbered': 0,
        'duplicates_removed': 0,
        'invalid_rows_skipped': 0,
    }

    def _fill(value, fallback):
        if value:
            return value
        repair_stats['filled_fields'] += 1
        return fallback

    def _text_value(value):
        if value is None:
            return ''
        if isinstance(value, list):
            return '\n'.join(str(item).strip() for item in value if str(item).strip())
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)
        return str(value).strip()

    def normalize_tc(d, index):
        scenario = _text_value(d.get('scenario') or d.get('test_scenario') or d.get('title') or d.get('name'))
        case_type_raw = _text_value(d.get('case_type') or d.get('type') or d.get('test_type') or 'Positive')
        precondition = _text_value(d.get('precondition') or d.get('pre_condition') or d.get('preconditions') or d.get('prerequisites'))
        steps = _text_value(d.get('steps') or d.get('step_scenario') or d.get('test_steps') or d.get('actions'))
        expected = _text_value(d.get('expected') or d.get('expected_result') or d.get('expected_results') or d.get('result'))

        case_type_map = {
            'positive': 'Positive',
            'negative': 'Negative',
            'boundary': 'Boundary',
        }
        case_type = case_type_map.get(case_type_raw.lower())
        if not case_type:
            case_type = 'Positive'
            repair_stats['case_type_fixed'] += 1

        scenario_label = scenario or f'{module_name} Generated Scenario {index}'
        scenario = _fill(scenario, f'{module_name} - Generated Scenario {index}')
        if ' - ' not in scenario:
            scenario = f'{module_name} - {scenario}'
            repair_stats['filled_fields'] += 1
        precondition = _fill(precondition, f'The {module_name} page is open and the required UI is available.')
        if not steps:
            steps = (
                f'1. Open {module_name}.\n'
                f'2. Execute the action described in {scenario_label}.\n'
                f'3. Verify the resulting UI or data state.'
            )
            repair_stats['filled_fields'] += 1
        elif not re.match(r'^\s*\d+\.', steps):
            step_lines = [line.strip() for line in re.split(r'\r?\n+', steps) if line.strip()]
            if not step_lines:
                step_lines = [steps]
            steps = '\n'.join(f'{line_no}. {line}' for line_no, line in enumerate(step_lines, 1))
            repair_stats['steps_numbered'] += 1
        expected = _fill(expected, f'The {module_name} flow completes successfully and shows the expected result for {scenario_label}.')

        return {
            'tc_id': d.get('tc_id') or d.get('id') or d.get('test_id') or '',
            'scenario': scenario,
            'case_type': case_type,
            'precondition': precondition,
            'steps': steps,
            'expected': expected,
        }

    test_cases_list = []
    seen_scenarios = set()
    for index, tc_data in enumerate(data.get('test_cases', []), 1):
        if not isinstance(tc_data, dict):
            repair_stats['invalid_rows_skipped'] += 1
            continue
        try:
            normalized_tc = normalize_tc(tc_data, index)
            scenario_key = normalized_tc['scenario'].strip().lower()
            if scenario_key in seen_scenarios:
                repair_stats['duplicates_removed'] += 1
                continue
            seen_scenarios.add(scenario_key)
            test_cases_list.append(TestCaseSchema(**normalized_tc))
        except Exception:
            repair_stats['invalid_rows_skipped'] += 1

    if not test_cases_list:
        raise Exception("AI returned no valid test cases. The model may have responded in an unexpected format. Try again or use a different model.")

    quality_warnings = evaluate_quality(test_cases_list, gen_depth)

    elements_list = []
    checklist_list = []
    if generate_script:
        for el_data in data.get('elements', []):
            try:
                elements_list.append(ElementSchema(**el_data))
            except Exception:
                pass
        for chk_data in data.get('checklist', []):
            try:
                checklist_list.append(ChecklistSchema(**chk_data))
            except Exception:
                pass

    clean_title = re.sub(r'[^a-zA-Z0-9_]', '', page_title.replace(' ', '_'))
    filename_base = f"{clean_title.lower()}"
    xlsx_name = f"testplan_{filename_base}.xlsx"

    outputs_dir = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'outputs')
    module_dir = os.path.join(outputs_dir, filename_base)
    os.makedirs(os.path.join(module_dir, "excel"), exist_ok=True)
    xlsx_path = os.path.join(module_dir, "excel", xlsx_name)

    if job_id:
        _set_job_state(job_id, message="Writing Excel workbook...", progress=82)
    tc_count, ordered_cases = build_excel_file(page_title, id_prefix, test_cases_list, elements_list, checklist_list, xlsx_path, model_name=model_name, gen_depth=gen_depth)

    py_name = None
    if generate_script:
        if job_id:
            _set_job_state(job_id, message="Writing optional Python script...", progress=92)
        os.makedirs(os.path.join(module_dir, "scripts"), exist_ok=True)
        py_name = f"testplan_{filename_base}.py"
        py_path = os.path.join(module_dir, "scripts", py_name)
        build_python_script(page_title, id_prefix, ordered_cases, elements_list, checklist_list, py_path, filename_base, model_name=model_name, gen_depth=gen_depth)

    checklist_target = 5 if gen_depth == "ultra" else (15 if gen_depth == "fast" else 50)
    preview_cases = [
        {
            'tc_id': tc.tc_id,
            'scenario': tc.scenario,
            'case_type': tc.case_type,
            'expected': tc.expected,
        }
        for tc in ordered_cases[:5]
    ]
    return {
        'test_case_count': tc_count,
        'checklist_count': checklist_target,
        'sheet_count': 2,
        'xlsx_file': xlsx_name,
        'py_file': py_name,
        'download_url': f'/api/export/{xlsx_name}/xlsx',
        'preview_url': f'/api/preview/{xlsx_name}',
        'provider': provider,
        'model_name': model_name,
        'generation_mode': gen_depth,
        'auto_repair_count': sum(repair_stats.values()),
        'repair_stats': repair_stats,
        'skipped_count': repair_stats['invalid_rows_skipped'] + repair_stats['duplicates_removed'],
        'quality_warnings': quality_warnings,
        'preview_cases': preview_cases,
        'generated_at': datetime.utcnow().isoformat(),
    }

# -----------------------------------------------------------------------------
# HELPER FUNCTIONS TO BUILD EXCEL FILE
# -----------------------------------------------------------------------------

def build_excel_file(title, prefix, test_cases, elements, checklist, output_path, model_name="gemini-3.5-flash", gen_depth="exhaustive"):
    wb = openpyxl.Workbook()
    
    # Fonts, fills, borders, alignments
    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    body_font    = Font(name="Calibri", size=11)
    wrap_align   = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center_align = Alignment(wrap_text=True, vertical="top", horizontal="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    def style_header(ws, row, max_col):
        for col in range(1, max_col + 1):
            c = ws.cell(row=row, column=col)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, center_align, thin_border
            
    def style_body(ws, row, max_col):
        for col in range(1, max_col + 1):
            c = ws.cell(row=row, column=col)
            c.font = body_font
            c.alignment = wrap_align if col != 1 else center_align
            c.border = thin_border

    def section_row(ws, row, sec_title, max_col):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        c = ws.cell(row=row, column=1, value=sec_title)
        c.font = Font(name="Calibri", bold=True, color="1F3864", size=12)
        c.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border
        for col in range(2, max_col + 1):
            ws.cell(row=row, column=col).border = thin_border

# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
    ws1 = wb.active
    sheet1_title = f"TEST PLAN - {title}"
    if len(sheet1_title) > 30:
        sheet1_title = sheet1_title[:27] + "..."
    ws1.title = sheet1_title
    
    headers1 = ["TC-ID", "TEST SCENARIO", "CASE TYPE", "PRE-CONDITION", "STEP SCENARIO", "EXPECTED RESULT", "ACTUAL RESULT", "EVIDENCE"]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers1))
    
    widths1 = [10, 45, 14, 30, 60, 55, 18, 18]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"
    
    # Sort test cases by area/component or keep order, and insert section rows
    # Group by Feature Area for section headers
    grouped_cases = {}
    for tc in test_cases:
        parts = tc.scenario.split(' - ')
        area = parts[0].strip().strip('[]') if len(parts) > 1 else "General"
        if area not in grouped_cases:
            grouped_cases[area] = []
        grouped_cases[area].append(tc)

# -----------------------------------------------------------------------------
    ordered_cases = [tc for cases in grouped_cases.values() for tc in cases]

    current_row = 2
    tc_counter = 1
    for area, cases in grouped_cases.items():
        section_row(ws1, current_row, f"SECTION: {area.upper()} TESTS", len(headers1))
        current_row += 1
        for tc in cases:
            ws1.cell(row=current_row, column=1, value=f"{prefix}{tc_counter:03d}")
            ws1.cell(row=current_row, column=2, value=tc.scenario)
            ws1.cell(row=current_row, column=3, value=tc.case_type)
            ws1.cell(row=current_row, column=4, value=tc.precondition)
            ws1.cell(row=current_row, column=5, value=tc.steps)
            ws1.cell(row=current_row, column=6, value=tc.expected)
            ws1.cell(row=current_row, column=7, value="") # Actual result empty
            ws1.cell(row=current_row, column=8, value="") # Evidence placeholder
            style_body(ws1, current_row, len(headers1))
            current_row += 1
            tc_counter += 1

# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
    ws2 = wb.create_sheet("SUMMARY")
    for c, h in enumerate(["METRIC", "VALUE"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center_align, thin_border
    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 25
    ws2.freeze_panes = "A2"
    
    _count_positive  = sum(1 for tc in test_cases if tc.case_type == "Positive")
    _count_negative  = sum(1 for tc in test_cases if tc.case_type == "Negative")
    _count_boundary  = sum(1 for tc in test_cases if tc.case_type == "Boundary")
    _count_total     = len(test_cases)

    def _kw_count(*keywords):
        return sum(
            1 for tc in test_cases
            if any(kw.lower() in tc.scenario.lower() for kw in keywords)
        )

    summary_rows = [
        ("TOTAL TEST CASES", _count_total),
        ("", ""),
        ("  Positive",  _count_positive),
        ("  Negative",  _count_negative),
        ("  Boundary",  _count_boundary),
        ("", ""),
        ("FORM/UI COMPONENT VALIDATION", ""),
    ]

    for area, cases in grouped_cases.items():
        summary_rows.append((f"  {area} Component Tests", len(cases)))

    summary_rows.extend([
        ("", ""),
        ("NON-FUNCTIONAL TESTING (ESTIMATED)", ""),
        ("  Accessibility (WCAG 2.1 AA)",    _kw_count("Accessibility", "WCAG", "Contrast")),
        ("  Cross-Browser Compatibility",    _kw_count("Browser", "Chrome", "Safari")),
        ("  Mobile Responsive Layout",       _kw_count("Mobile", "Responsive", "Viewport")),
        ("  Performance & Latency",          _kw_count("Performance", "Slow 3G", "Timeout")),
        ("  Security & Sanitization",        _kw_count("Security", "XSS", "Injection")),
        ("", ""),
        ("TEST PLAN LANGUAGE", "English"),
        ("GENERATION ENGINE", f"{model_name} (Himagent AI)")
    ])
    
    for idx, (metric, value) in enumerate(summary_rows, 2):
        c1 = ws2.cell(row=idx, column=1, value=metric)
        c2 = ws2.cell(row=idx, column=2, value=value)
        c1.font = Font(name="Calibri", bold=(not metric.startswith("  ")), size=11) if metric else body_font
        c2.font = body_font
        for c in (c1, c2):
            c.border = thin_border
            c.alignment = wrap_align

    # Add programmatic Pie Chart to SUMMARY sheet (distributing Case Types)
    try:
        from openpyxl.chart import PieChart, Reference
        chart = PieChart()
        labels = Reference(ws2, min_col=1, min_row=4, max_row=6)
        data = Reference(ws2, min_col=2, min_row=4, max_row=6)
        chart.add_data(data)
        chart.set_categories(labels)
        chart.title = "Case Type Distribution"
        chart.width = 14
        chart.height = 7
        ws2.add_chart(chart, "D2")
    except Exception as e:
        print(f"Error adding chart to Excel: {e}")

# -----------------------------------------------------------------------------
    # CONDITIONAL FORMATTING
# -----------------------------------------------------------------------------
    green_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill    = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Case type formatting on Sheet 1 (Column C)
    ws1.conditional_formatting.add(f"C2:C{current_row - 1}", CellIsRule(operator='equal', formula=['"Positive"'], fill=green_fill))
    ws1.conditional_formatting.add(f"C2:C{current_row - 1}", CellIsRule(operator='equal', formula=['"Negative"'], fill=red_fill))
    ws1.conditional_formatting.add(f"C2:C{current_row - 1}", CellIsRule(operator='equal', formula=['"Boundary"'], fill=yellow_fill))

    wb.save(output_path)
    return current_row - 2, ordered_cases

# -----------------------------------------------------------------------------
# HELPER TO GENERATE SELF-CONTAINED PYTHON SCRIPT
# -----------------------------------------------------------------------------

def build_python_script(title, prefix, test_cases, elements, checklist, output_path, filename_base, model_name="gemini-3.5-flash", gen_depth="exhaustive"):
    """Generate a thin Python script that imports from core/ modules to recreate the workbook."""
    # Prepare serializable data
    tc_tuples = []
    for idx, tc in enumerate(test_cases, start=1):
        tc_tuples.append((
            f"{prefix}{idx:03d}",
            tc.scenario,
            tc.case_type,
            tc.precondition,
            tc.steps,
            tc.expected
        ))

    el_tuples = []
    for el in elements:
        el_tuples.append((el.area, el.element_name, el.element_type, el.interactions))

    checklist_target = 5 if gen_depth == "ultra" else (15 if gen_depth == "fast" else 50)
    final_checklist = checklist[:checklist_target]
    while len(final_checklist) < checklist_target:
        pad_index = len(final_checklist) + 1
        final_checklist.append(ChecklistSchema(
            category="Security",
            checklist_item=f"Verify that form validation security filter {pad_index} checks against common cross-site scripting vulnerabilities"
        ))

    chk_tuples = []
    for idx, chk in enumerate(final_checklist, 1):
        chk_tuples.append((
            str(idx),
            chk.checklist_item,
            chk.category,
            "Manual" if chk.category in ["UI/UX Layout", "Accessibility (WCAG)", "Mobile Responsive"] else "Automated",
            "PENDING"
        ))

    sheet1_title = f"TEST PLAN - {title}"[:30]

    # Build summary rows dynamically (group by area from scenario prefix)
    grouped_areas = {}
    for tc in test_cases:
        parts = tc.scenario.split(' - ')
        area = parts[0].strip().strip('[]') if len(parts) > 1 else "General"
        grouped_areas[area] = grouped_areas.get(area, 0) + 1

    summary_rows_repr = [
        ("TOTAL TEST CASES", "=SUM(B4:B6)"),
        ("", ""),
        ("  Positive", f"=COUNTIF('{sheet1_title}'!C:C, \"Positive\")"),
        ("  Negative", f"=COUNTIF('{sheet1_title}'!C:C, \"Negative\")"),
        ("  Boundary", f"=COUNTIF('{sheet1_title}'!C:C, \"Boundary\")"),
        ("", ""),
        ("FORM/UI COMPONENT VALIDATION", ""),
    ]
    for area in grouped_areas:
        summary_rows_repr.append((f"  {area} Component Tests", f"=COUNTIF('{sheet1_title}'!B:B, \"*{area}*\")"))
    summary_rows_repr.extend([
        ("", ""),
        ("NON-FUNCTIONAL TESTING (ESTIMATED)", ""),
        ("  Accessibility (WCAG 2.1 AA)", f"=COUNTIF('{sheet1_title}'!B:B, \"*Accessibility*\") + COUNTIF('{sheet1_title}'!B:B, \"*WCAG*\")"),
        ("  Cross-Browser Compatibility",   f"=COUNTIF('{sheet1_title}'!B:B, \"*Browser*\") + COUNTIF('{sheet1_title}'!B:B, \"*Chrome*\")"),
        ("  Mobile Responsive Layout",       f"=COUNTIF('{sheet1_title}'!B:B, \"*Mobile*\") + COUNTIF('{sheet1_title}'!B:B, \"*Responsive*\")"),
        ("  Performance & Latency",          f"=COUNTIF('{sheet1_title}'!B:B, \"*Performance*\") + COUNTIF('{sheet1_title}'!B:B, \"*Slow 3G*\")"),
        ("  Security & Sanitization",        f"=COUNTIF('{sheet1_title}'!B:B, \"*Security*\") + COUNTIF('{sheet1_title}'!B:B, \"*XSS*\")"),
        ("", ""),
        ("TEST PLAN LANGUAGE",  "English"),
        ("GENERATION ENGINE",   f"{model_name} (Himagent AI)"),
    ])

    HEADERS    = ["TC-ID", "TEST SCENARIO", "CASE TYPE", "PRE-CONDITION", "STEP SCENARIO", "EXPECTED RESULT", "ACTUAL RESULT", "EVIDENCE"]
    COL_WIDTHS = [10, 45, 14, 30, 60, 55, 18, 18]

    script_content = f'''import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from core.sheets import (
    build_testcase_sheet_v2, build_summary_sheet, apply_conditional_fmt
)

# -----------------------------------------------------------------------------
SHEET_TITLE = {repr(sheet1_title)}
HEADERS     = {repr(HEADERS)}
COL_WIDTHS  = {repr(COL_WIDTHS)}

# test_cases: (tc_id, scenario, case_type, precondition, steps, expected)
# Wrapped as (tc_id, scenario_key, test_name, case_type, precond, steps, expected)
# where scenario_key uses the tc_id as area so sections group by tc_id prefix
TEST_CASES = {repr(tc_tuples)}

# Convert flat (tc_id, scenario, case_type, precond, steps, expected)
# into the format expected by build_testcase_sheet_v2:
# (tc_id, scenario_code, test_name, case_type, precond, steps, expected)
def _to_v2_format(rows):
    out = []
    for tc_id, scenario, case_type, precond, steps, expected in rows:
        parts = scenario.split(" - ", 1)
        area_code = parts[0].strip().replace(" ", "_").upper() if len(parts) > 1 else "GENERAL"
        test_name = parts[1].strip() if len(parts) > 1 else scenario
        out.append((tc_id, area_code, test_name, case_type, precond, steps, expected))
    return out

SUMMARY    = {repr(summary_rows_repr)}

# -----------------------------------------------------------------------------
wb = openpyxl.Workbook()

ws_tests, last_row = build_testcase_sheet_v2(
    wb, SHEET_TITLE, HEADERS, COL_WIDTHS, _to_v2_format(TEST_CASES)
)
ws_summary   = build_summary_sheet(wb, SHEET_TITLE, SUMMARY)
apply_conditional_fmt(
    ws_tests,     f"C2:C{{last_row - 1}}"
)

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), {repr(f"testplan_{filename_base}.xlsx")})
wb.save(OUTPUT)
print(f"Saved: {{OUTPUT}}")
print(f"Test cases: {{len(TEST_CASES)}}")
'''
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(script_content)

# -----------------------------------------------------------------------------
# FLASK WEB ENDPOINTS
# -----------------------------------------------------------------------------

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/router/models', methods=['GET'])
def router_models():
    try:
        models = fetch_router_models()
    except Exception as exc:
        return jsonify({
            'success': False,
            'message': format_router_ping_error(exc, ROUTER_DEFAULT_MODEL),
            'models': [],
        }), 503
    if not models:
        return jsonify({
            'success': False,
            'message': '9Router returned no models.',
            'models': [],
        }), 503
    return jsonify({
        'success': True,
        'default_model': models[0]['value'],
        'models': models,
    })

@app.route('/api/ping', methods=['POST'])
def ping_api_key():
    """Test if an API key is valid by making a minimal request to the provider."""
    provider = ''
    model_name = ''
    try:
        provider   = detect_provider(request.form.get('provider',''), request.form.get('model_name',''), '')
        api_key    = request.form.get('api_key', '').strip()
        model_name = request.form.get('model_name', '').strip()
        model_name = normalize_router_model(provider, model_name)
        if not api_key and provider != '9router':
            return jsonify({'success': False, 'message': 'No API key provided.'})

        if provider == '9router':
            result, status_code = test_router_model_route(model_name)
            return jsonify(result), status_code

        if provider == 'gemini':
            genai.configure(api_key=api_key)
            m = genai.GenerativeModel(model_name or 'gemini-2.5-flash')
            m.generate_content("hi", generation_config=genai.GenerationConfig(max_output_tokens=1))

        elif provider == 'claude':
            r = requests.post('https://api.anthropic.com/v1/messages',
                headers={'x-api-key': api_key, 'anthropic-version': '2023-06-01', 'content-type': 'application/json'},
                json={'model': model_name or 'claude-haiku-4-5', 'max_tokens': 1,
                      'messages': [{'role': 'user', 'content': 'hi'}]}, timeout=15)
            if r.status_code != 200:
                return jsonify({'success': False, 'message': parse_api_error(r.status_code, r.text, 'Claude')})

        else:
            base_urls = {
                'openai':   'https://api.openai.com/v1',
                'deepseek': 'https://api.deepseek.com',
                'grok':     'https://api.x.ai/v1',
                'mistral':  'https://api.mistral.ai/v1',
                'mimo':     'https://api.xiaomimimo.com/v1',
            }
            base_url = base_urls.get(provider, 'https://api.openai.com/v1')
            try:
                r = requests.post(f'{base_url}/chat/completions',
                    headers=({'Content-Type': 'application/json'} if not api_key else {'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'}),
                    json={'model': model_name, 'max_tokens': 1,
                          'messages': [{'role': 'user', 'content': 'hi'}]}, timeout=15)
            except requests.RequestException as exc:
                if provider == '9router':
                    return jsonify({'success': False, 'message': format_router_ping_error(exc, model_name)})
                raise
            if r.status_code != 200:
                if provider == '9router' and r.status_code in (401, 403, 404):
                    return jsonify({
                        'success': False,
                        'message': (
                            f"9Router route/model '{model_name}' is not ready ({r.status_code}). "
                            "Open http://localhost:20128, make sure that route is logged in, enabled, and mapped to a working upstream provider."
                        )
                    })
                return jsonify({'success': False, 'message': parse_api_error(r.status_code, r.text, provider.title())})

        return jsonify({'success': True})
    except Exception as e:
        if google_exceptions and isinstance(e, (
            google_exceptions.ResourceExhausted, google_exceptions.PermissionDenied,
            google_exceptions.Unauthenticated, google_exceptions.NotFound,
            google_exceptions.InvalidArgument, google_exceptions.ServiceUnavailable
        )):
            return jsonify({'success': False, 'message': parse_gemini_exception(e)})
        if provider == '9router':
            return jsonify({'success': False, 'message': format_router_ping_error(e, model_name)})
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/generate', methods=['POST'])
def generate_test_plan():
    try:
        # Check files
        uploaded_files = request.files.getlist('screenshot')
        if not uploaded_files or all(f.filename == '' for f in uploaded_files):
            return jsonify({'success': False, 'message': 'No screenshot files uploaded.'}), 400
            
        page_title = request.form.get('page_title', 'Custom SUT').strip()
        id_prefix = request.form.get('id_prefix', 'TC-C-').strip()
        model_name = request.form.get('model_name', 'gemini-3.5-flash').strip()
        instructions = request.form.get('instructions', '').strip()
        screen_contexts = [
            str(value or '').strip()[:280]
            for value in request.form.getlist('screen_context')
        ]
        user_api_key = request.form.get('api_key', '').strip()
        req_provider = request.form.get('provider', '').strip()
        gen_depth = request.form.get('gen_depth', 'fast').strip().lower()
        if gen_depth not in ['ultra', 'fast', 'exhaustive']:
            gen_depth = 'fast'
        generate_script = request.form.get('generate_script', '0').strip().lower() in ('1', 'true', 'yes', 'on')

        files_payload = []
        for f in uploaded_files:
            if f.filename:
                files_payload.append({'filename': f.filename, 'bytes': f.read()})
        if not files_payload:
            return jsonify({'success': False, 'message': 'No valid screenshot files uploaded.'}), 400
        try:
            validate_screenshot_payload(files_payload)
        except ValueError as validation_error:
            return jsonify({'success': False, 'message': str(validation_error)}), 400

        payload = {
            'files': files_payload,
            'page_title': page_title,
            'id_prefix': id_prefix,
            'model_name': model_name,
            'instructions': instructions,
            'screen_contexts': screen_contexts,
            'api_key': user_api_key,
            'provider': req_provider,
            'gen_depth': gen_depth,
            'generate_script': generate_script,
        }
        cache_key = _make_generation_cache_key(payload)
        with _GENERATION_LOCK:
            cached = _GENERATION_CACHE.get(cache_key)
        if not cached:
            cached = get_generation_cache(_GENERATION_DB_PATH, cache_key)
        if cached:
            job_id = str(uuid4())
            _set_job_state(
                job_id,
                status='completed',
                message='Completed from cache.',
                progress=100,
                result=cached,
                cached=True,
                cache_key=cache_key,
            )
            return jsonify({
                'success': True,
                'queued': False,
                'cached': True,
                'job_id': job_id,
                'status_url': f'/api/jobs/{job_id}',
                'message': 'Result loaded from cache.',
                'progress': 100,
                **cached,
            }), 200

        job_id = str(uuid4())
        with _GENERATION_LOCK:
            active_job_id = _GENERATION_ACTIVE_BY_CACHE.get(cache_key)
            active_job = _GENERATION_JOBS.get(active_job_id, {}) if active_job_id else {}
            if active_job_id and active_job.get('status') in ('queued', 'running'):
                return jsonify({
                    'success': True,
                    'queued': True,
                    'job_id': active_job_id,
                    'status_url': f'/api/jobs/{active_job_id}',
                    'message': 'Generation already running for this request.',
                    'deduped': True,
                }), 202
            _GENERATION_ACTIVE_BY_CACHE[cache_key] = job_id

        _set_job_state(
            job_id,
            status='queued',
            message='Queued for generation...',
            progress=0,
            result=None,
            error=None,
            cached=False,
            cache_key=cache_key,
        )

        def _worker():
            try:
                result = _run_generation_pipeline(payload, job_id=job_id)
                with _GENERATION_LOCK:
                    _GENERATION_CACHE[cache_key] = result
                set_generation_cache(_GENERATION_DB_PATH, cache_key, result)
                _set_job_state(
                    job_id,
                    status='completed',
                    message='Generation complete.',
                    progress=100,
                    result=result,
                    error=None,
                )
            except Exception as e:
                _set_job_state(
                    job_id,
                    status='failed',
                    message=str(e),
                    error=str(e),
                    progress=100,
                )
            finally:
                with _GENERATION_LOCK:
                    if _GENERATION_ACTIVE_BY_CACHE.get(cache_key) == job_id:
                        _GENERATION_ACTIVE_BY_CACHE.pop(cache_key, None)

        threading.Thread(target=_worker, daemon=True).start()
        return jsonify({
            'success': True,
            'queued': True,
            'job_id': job_id,
            'status_url': f'/api/jobs/{job_id}',
            'message': 'Generation queued.',
        }), 202
            
    except Exception as e:
        traceback.print_exc()
        if google_exceptions and isinstance(e, (
            google_exceptions.ResourceExhausted, google_exceptions.PermissionDenied,
            google_exceptions.Unauthenticated, google_exceptions.NotFound,
            google_exceptions.InvalidArgument, google_exceptions.ServiceUnavailable
        )):
            msg = parse_gemini_exception(e)
        else:
            msg = str(e)
        return jsonify({'success': False, 'message': msg}), 500


@app.route('/api/jobs/<job_id>', methods=['GET'])
def get_generation_job(job_id):
    job = _get_job_state(job_id)
    if not job:
        return jsonify({'success': False, 'message': 'Job not found.'}), 404
    response = {
        'success': True,
        'job_id': job_id,
        'status': job.get('status', 'unknown'),
        'message': job.get('message', ''),
        'progress': job.get('progress', 0),
        'cached': job.get('cached', False),
    }
    if job.get('status') == 'completed' and job.get('result'):
        response.update(job['result'])
    if job.get('status') == 'failed':
        response['error'] = job.get('error') or job.get('message') or 'Generation failed.'
    return jsonify(response)

@app.route('/api/model-health', methods=['GET'])
def model_health():
    """Return local model health from previous generation jobs only."""
    try:
        rows = []
        if os.path.exists(_GENERATION_DB_PATH):
            with sqlite3.connect(_GENERATION_DB_PATH) as conn:
                conn.row_factory = sqlite3.Row
                rows = conn.execute(
                    """
                    SELECT status, result_json, updated_at
                    FROM generation_jobs
                    WHERE result_json IS NOT NULL
                    ORDER BY updated_at DESC
                    LIMIT 200
                    """
                ).fetchall()
        health = {}
        for row in rows:
            try:
                result = json.loads(row['result_json'] or '{}')
            except Exception:
                continue
            provider = result.get('provider') or ''
            model_name = result.get('model_name') or ''
            if not model_name:
                continue
            key = f"{provider}:{model_name}"
            item = health.setdefault(key, {
                'provider': provider,
                'model_name': model_name,
                'success_count': 0,
                'failure_count': 0,
                'last_status': row['status'],
                'last_seen': row['updated_at'],
            })
            if row['status'] == 'completed':
                item['success_count'] += 1
            elif row['status'] == 'failed':
                item['failure_count'] += 1
            if row['updated_at'] and row['updated_at'] > (item.get('last_seen') or ''):
                item['last_status'] = row['status']
                item['last_seen'] = row['updated_at']
        return jsonify({'success': True, 'models': list(health.values())})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc), 'models': []}), 500

def find_file_in_outputs(filename):
    """Search outputs/ recursively for a file matching filename. Returns full path or None."""
    return core_find_file_in_outputs(filename)

@app.route('/api/download/<filename>')
def download_file(filename):
    filename = safe_filename(filename)
    if filename.endswith('.xlsx'):
        return export_file(filename, 'xlsx')
    file_path = find_file_in_outputs(filename)
    if file_path and os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return "File not found", 404

# New endpoint to delete a generated file locally
@app.route('/api/delete/<filename>', methods=['DELETE'])
def delete_file(filename):
    """Delete a file from the outputs directory safely."""
    # Sanitize filename to prevent path traversal
    filename = safe_filename(filename)
    if not find_file_in_outputs(filename):
        return jsonify({'success': False, 'message': f'File {filename} not found.'}), 404
    try:
        core_delete_generated_file(filename)
        return jsonify({'success': True, 'message': f'File {filename} deleted.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# -----------------------------------------------------------------------------
# NEW API ENDPOINTS: LIBRARY LISTING & FILE PREVIEW
# -----------------------------------------------------------------------------
@app.route('/api/files', methods=['GET'])
def list_files():
    try:
        tree = list_generated_tree()
        return jsonify({'success': True, 'tree': tree})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/preview/<filename>', methods=['GET'])
def get_preview(filename):
    try:
        filename = secure_filename(filename)
        file_path = find_file_in_outputs(filename)
        if not file_path or not os.path.exists(file_path):
            return jsonify({'success': False, 'message': f'File {filename} not found.'}), 404
        result = get_file_preview(file_path, filename)
        if not result.get('success'):
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
@app.route('/api/save_xlsx', methods=['POST'])
def save_xlsx():
    try:
        payload = request.get_json(force=True)
        filename = safe_filename(payload.get('filename', ''))
        changes = payload.get('changes', {}) # dict: {sheet_name: [{'coordinate': 'A1', 'value': 'val'}, ...]}
        
        file_path = find_file_in_outputs(filename)
        if not file_path or not os.path.exists(file_path):
            return jsonify({'success': False, 'message': f'File {filename} not found.'}), 404

        wb = openpyxl.load_workbook(file_path)
        for sheet_name, sheet_changes in changes.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for chg in sheet_changes:
                    coord = chg.get('coordinate')
                    val = chg.get('value')
                    if coord:
                        ws[coord] = val
        wb.save(file_path)

        return jsonify({'success': True, 'message': 'Changes saved successfully!'})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# MULTI-FORMAT EXPORT SYSTEM
# -----------------------------------------------------------------------------
@app.route('/api/export/<filename>/<export_format>', methods=['GET'])
def export_file(filename, export_format):
    try:
        filename = safe_filename(filename)
        export_format = export_format.lower().strip()
        file_path = find_file_in_outputs(filename)
        if not file_path or not os.path.exists(file_path) or not filename.endswith('.xlsx'):
            return "File not found or invalid format", 404
        data, mimetype, download_name = export_workbook(file_path, filename, export_format)
        if data is None:
            return "Unsupported format", 400
        return send_file(data, mimetype=mimetype, as_attachment=True, download_name=download_name)
    except Exception as e:
        traceback.print_exc()
        return str(e), 500

# -----------------------------------------------------------------------------
# AUTO-DETECT MODULE TYPE + AI NAMING SUGGESTION
# -----------------------------------------------------------------------------
@app.route('/api/detect', methods=['POST'])
def detect_module():
    """Lightweight call: detect module type + suggest name & TC prefix from screenshot."""
    try:
        uploaded_files = request.files.getlist('screenshot')
        if not uploaded_files or all(f.filename == '' for f in uploaded_files):
            return jsonify({'success': False, 'message': 'No screenshot provided.'}), 400

        user_api_key = request.form.get('api_key', '').strip()
        model_name = request.form.get('model_name', 'gemini-3.5-flash').strip()
        req_provider = request.form.get('provider', '').strip()

        provider = detect_provider(req_provider, model_name, user_api_key)
        model_name = normalize_router_model(provider, model_name)
        api_key = resolve_api_key(provider, user_api_key)

        if not api_key and provider != '9router':
            return jsonify({'success': False, 'message': f'{provider.upper()} API key required.'}), 400

        first_file = next((f for f in uploaded_files if f.filename != ''), None)
        if not first_file:
            return jsonify({'success': False, 'message': 'No valid screenshot.'}), 400

        raw_bytes = first_file.read()
        try:
            validate_screenshot_payload([{'filename': first_file.filename, 'bytes': raw_bytes}])
        except ValueError as validation_error:
            return jsonify({'success': False, 'message': str(validation_error)}), 400

        pil_img = Image.open(BytesIO(raw_bytes))
        result = detect_module_from_screenshot(provider, model_name, api_key, pil_img)
        return jsonify({'success': True, **result})

    except Exception as e:
        traceback.print_exc()
        if google_exceptions and isinstance(e, (
            google_exceptions.ResourceExhausted, google_exceptions.PermissionDenied,
            google_exceptions.Unauthenticated, google_exceptions.NotFound,
            google_exceptions.InvalidArgument, google_exceptions.ServiceUnavailable
        )):
            msg = parse_gemini_exception(e)
        else:
            msg = str(e)
        return jsonify({'success': False, 'message': msg}), 500


# -----------------------------------------------------------------------------
# DASHBOARD STATISTICS API
# -----------------------------------------------------------------------------
@app.route('/api/stats', methods=['GET'])
def get_stats():
    """Return aggregated statistics from the outputs/ folder."""
    try:
        outputs_dir = _outputs_dir()
        os.makedirs(outputs_dir, exist_ok=True)

        file_map = _build_outputs_file_map()
        total_plans = 0
        total_test_cases = 0
        total_size_bytes = 0
        files_by_date = {}
        recent_files = []
        seen_xlsx = set()

        def add_output_file(file_name, file_path, test_case_count=None):
            nonlocal total_plans, total_test_cases, total_size_bytes
            if not file_path or not os.path.exists(file_path) or file_name in seen_xlsx:
                return

            stat = os.stat(file_path)
            total_plans += 1
            total_size_bytes += stat.st_size
            date_str = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d')
            files_by_date[date_str] = files_by_date.get(date_str, 0) + 1
            if test_case_count is None:
                test_case_count = _count_test_cases_from_workbook(file_path)
            try:
                total_test_cases += int(test_case_count or 0)
            except Exception:
                pass

            recent_files.append({
                'name': file_name,
                'size': stat.st_size,
                'modified': stat.st_mtime,
                'module': _filename_to_module_label(file_name),
                'type': 'excel',
            })
            seen_xlsx.add(file_name)

        if os.path.exists(_GENERATION_DB_PATH):
            with sqlite3.connect(_GENERATION_DB_PATH) as conn:
                conn.row_factory = sqlite3.Row
                rows = conn.execute(
                    """
                    SELECT result_json, created_at, updated_at
                    FROM generation_jobs
                    WHERE status = 'completed' AND result_json IS NOT NULL
                    ORDER BY COALESCE(updated_at, created_at) DESC
                    """
                ).fetchall()

            for row in rows:
                try:
                    result = json.loads(row['result_json'])
                except Exception:
                    continue
                xlsx_name = result.get('xlsx_file')
                if not xlsx_name:
                    continue
                add_output_file(
                    xlsx_name,
                    file_map.get(xlsx_name),
                    result.get('test_case_count'),
                )

        for file_name, file_path in file_map.items():
            if not file_name.endswith('.xlsx'):
                continue
            add_output_file(file_name, file_path)

        recent_files.sort(key=lambda item: item['modified'], reverse=True)
        sorted_dates = sorted(files_by_date.keys())
        chart_labels = sorted_dates[-14:]
        chart_values = [files_by_date.get(date_key, 0) for date_key in chart_labels]

        return jsonify({
            'success': True,
            'total_plans': total_plans,
            'total_test_cases': total_test_cases,
            'total_size_bytes': total_size_bytes,
            'total_size_kb': round(total_size_bytes / 1024, 1),
            'chart_labels': chart_labels,
            'chart_values': chart_values,
            'recent_files': recent_files[:15],
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# -----------------------------------------------------------------------------
# CUSTOM PRESETS (server-side, data/custom_presets.json)
# -----------------------------------------------------------------------------
_PRESETS_FILE = os.path.join(os.path.abspath(os.path.dirname(__file__)), "data", "custom_presets.json")

def _load_presets():
    if not os.path.exists(_PRESETS_FILE):
        return []
    try:
        with open(_PRESETS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def _save_presets(presets):
    os.makedirs(os.path.dirname(_PRESETS_FILE), exist_ok=True)
    with open(_PRESETS_FILE, "w", encoding="utf-8") as f:
        json.dump(presets, f, ensure_ascii=False, indent=2)

@app.route("/api/presets", methods=["GET"])
def get_presets():
    return jsonify(_load_presets())

@app.route("/api/presets", methods=["POST"])
def add_preset():
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    text = (data.get("text") or "").strip()
    category = (data.get("category") or "my").strip()
    if not name or not text:
        return jsonify({"error": "name and text required"}), 400
    presets = _load_presets()
    preset = {"id": str(uuid4()), "name": name, "text": text, "category": category}
    presets.append(preset)
    _save_presets(presets)
    return jsonify(preset), 201

@app.route("/api/presets/<preset_id>", methods=["DELETE"])
def delete_preset(preset_id):
    presets = [p for p in _load_presets() if p.get("id") != preset_id]
    _save_presets(presets)
    return jsonify({"ok": True})

# -----------------------------------------------------------------------------
if __name__ == '__main__':
    app.run(host='localhost', port=5000, debug=True, threaded=True, use_reloader=False)
