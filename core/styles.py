"""
core/styles.py
Centralized openpyxl style definitions shared across all test plan generators.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Header ──────────────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

# ── Body ─────────────────────────────────────────────────────────────────────
body_font = Font(name="Calibri", size=11)
sub_fill  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# ── Section row ──────────────────────────────────────────────────────────────
section_font = Font(name="Calibri", bold=True, color="1F3864", size=12)
section_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

# ── Alignment ────────────────────────────────────────────────────────────────
wrap_align   = Alignment(wrap_text=True, vertical="top", horizontal="left")
center_align = Alignment(wrap_text=True, vertical="top", horizontal="center")
section_align = Alignment(horizontal="left", vertical="center")

# ── Border ───────────────────────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

# ── Conditional Formatting Fills ─────────────────────────────────────────────
green_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
red_fill    = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
gray_fill   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
