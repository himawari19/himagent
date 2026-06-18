"""
core/sheets.py
Generic sheet-writing functions used by all test plan entry points.

  - build_main_testplan  : main pipeline builder (Sheet 1 + SUMMARY)
  - build_testcase_sheet_v2: Sheet 1 builder used by generated .py scripts
  - build_summary_sheet  : Sheet 2 - SUMMARY with pie chart
  - apply_conditional_fmt: conditional formatting on Case Type column
"""
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from core.styles import (
    header_font, header_fill, body_font,
    wrap_align, center_align, thin_border,
    green_fill, red_fill, yellow_fill, gray_fill
)
from core.helpers import style_header, style_body, section_row
from core.validators import add_casetype_validation


# ── Sheet 1 builder (used by generated .py scripts) ─────────────────────────

def build_testcase_sheet_v2(wb, sheet_title: str, headers: list,
                             col_widths: list, test_cases: list) -> tuple:
    """
    Build test-case sheet for imageGen/videoGen format.
    test_cases items: (scenario, test_name, case_type, precond, steps, expected)
      where scenario is the dotted area code e.g. 'NAVBAR.LOGO'

    Returns (ws, last_row).
    """
    ws = wb.active
    ws.title = sheet_title

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Group test cases into sections by scenario prefix
    current_section = None
    row = 2
    tc_counter = 0

    for item in test_cases:
        if len(item) == 7:
            tc_id, scenario, test_name, case_type, precond, steps, expected = item
        else:
            scenario, test_name, case_type, precond, steps, expected = item
            tc_counter += 1
            prefix_str = "TC-"
            if "IMAGE" in sheet_title.upper():
                prefix_str = "TC-I-"
            elif "VIDEO" in sheet_title.upper():
                prefix_str = "TC-V-"
            tc_id = f"{prefix_str}{tc_counter:03d}"

        # Detect section change (first dot-segment = area)
        area = scenario.split(".")[0]
        if area != current_section:
            current_section = area
            section_title = _area_to_section_title(area)
            section_row(ws, row, section_title, len(headers))
            row += 1

        scenario_display = f"{scenario} — {test_name}"
        # Align with headers: TC-ID, TEST SCENARIO, CASE TYPE, PRE-CONDITION, STEP SCENARIO, EXPECTED RESULT, ACTUAL RESULT, EVIDENCE
        values = [tc_id, scenario_display, case_type, precond, steps, expected, "", ""]
        for c, v in enumerate(values, 1):
            ws.cell(row=row, column=c, value=v)
        style_body(ws, row, len(headers))
        row += 1

    add_casetype_validation(ws, f"C2:C{row - 1}")
    return ws, row


def _area_to_section_title(area: str) -> str:
    """Convert area code like 'NAVBAR' or 'IMAGE_TALENT' to a section title."""
    mapping = {
        "NAVBAR": "A. NAVBAR ELEMENTS",
        "HEADER": "B. HEADER ELEMENTS",
        "MODEL": "C. MODEL SELECTION",
        "IMAGE_TALENT": "D. IMAGE TALENT",
        "IMAGE_PRODUCT": "E. IMAGE PRODUCT",
        "IMAGE_LOGO": "F. IMAGE LOGO",
        "ADVANCED": "G. ADVANCED OPTIONS",
        "PROMPT_LIBRARY": "H. PROMPT LIBRARY",
        "PROMPT_FIELD": "I. PROMPT FIELD",
        "GENERATE": "J. GENERATE BUTTON",
        "TOGGLES": "K. TOGGLES",
        "E2E": "L. E2E INTEGRATION FLOWS",
        "POST_GENERATE": "M. POST-GENERATE ACTIONS",
        "ERROR": "N. ERROR HANDLING & NETWORK",
        "UIUX": "O. UI/UX BEHAVIOR",
        "CROSS_BROWSER": "P. CROSS-BROWSER",
        "MOBILE": "Q. MOBILE & RESPONSIVE",
        "SESSION": "R. SESSION AUTH & CSRF",
        "PERFORMANCE": "S. PERFORMANCE",
        "I18N": "T. LOCALIZATION (i18n)",
        "ADVANCED_COMBO": "U. ASPECT RATIO × RESOLUTION COMBOS",
        "SECURITY": "V. SECURITY DEEP-DIVE",
        # Video-specific
        "FRAME": "C. IMAGE FRAMES UPLOAD",
        "DURATION": "E. DURATION SELECTOR",
        "ASPECT": "F. ASPECT RATIO SELECTOR",
        "RESOLUTION": "G. RESOLUTION SELECTOR",
        "AUDIO": "H. AUDIO SETTINGS",
        "RESULT": "M. RESULT AREA ACTIONS",
    }
    return mapping.get(area, f"{area.replace('_', ' ').title()} Tests")


# ── Sheet 2: SUMMARY ─────────────────────────────────────────────────────────

def build_summary_sheet(wb, sheet1_title: str, summary_rows: list):
    """
    Build the SUMMARY sheet.
    summary_rows: list of (metric_label, value_or_formula) tuples.
    """
    ws2 = wb.create_sheet("SUMMARY")
    for c, h in enumerate(["METRIC", "VALUE"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font, cell.fill = header_font, header_fill
        cell.alignment, cell.border = center_align, thin_border

    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 22
    ws2.freeze_panes = "A2"

    from openpyxl.styles import Font as _Font, PatternFill as _Fill
    for i, (metric, value) in enumerate(summary_rows, 2):
        c1 = ws2.cell(row=i, column=1, value=metric)
        c2 = ws2.cell(row=i, column=2, value=value)

        is_main_cat = metric and not metric.startswith("  ") and value == ""
        is_title    = metric == "TOTAL TEST CASES"

        c1.font = _Font(name="Calibri", bold=(is_main_cat or is_title), size=11)
        c2.font = _Font(name="Calibri", bold=is_title, size=11)

        for c in (c1, c2):
            c.border    = thin_border
            c.alignment = wrap_align if c.column == 1 else center_align
            if is_main_cat:
                c.fill = gray_fill

    # Pie chart
    try:
        from openpyxl.chart import PieChart, Reference
        chart = PieChart()
        labels = Reference(ws2, min_col=1, min_row=4, max_row=6)
        data   = Reference(ws2, min_col=2, min_row=4, max_row=6)
        chart.add_data(data)
        chart.set_categories(labels)
        chart.title  = "Case Type Distribution"
        chart.width  = 14
        chart.height = 7
        ws2.add_chart(chart, "D2")
    except Exception as e:
        print(f"[WARNING] Could not add chart to SUMMARY: {e}")

    return ws2


# ── Main testplan builder (used by _run_generation_pipeline) ─────────────────

def build_main_testplan(wb, title: str, prefix: str, test_cases: list,
                        model_name: str = "gemini", gen_depth: str = "exhaustive") -> tuple:
    """
    Build Sheet 1 (test cases) + Sheet 2 (SUMMARY) from Pydantic TestCaseSchema objects.
    Returns (tc_count, ordered_cases).
    """
    ws1 = wb.active
    sheet1_title = f"TEST PLAN - {title}"
    if len(sheet1_title) > 30:
        sheet1_title = sheet1_title[:27] + "..."
    ws1.title = sheet1_title

    headers = ["TC-ID", "TEST SCENARIO", "CASE TYPE", "PRE-CONDITION",
               "STEP SCENARIO", "EXPECTED RESULT", "ACTUAL RESULT", "EVIDENCE"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers))

    for i, w in enumerate([10, 45, 14, 30, 60, 55, 18, 18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # Group by feature area for section headers
    grouped: dict = {}
    for tc in test_cases:
        parts = tc.scenario.split(' - ')
        area = parts[0].strip().strip('[]') if len(parts) > 1 else "General"
        grouped.setdefault(area, []).append(tc)

    ordered_cases = [tc for cases in grouped.values() for tc in cases]
    current_row = 2
    tc_counter = 1
    for area, cases in grouped.items():
        section_row(ws1, current_row, f"SECTION: {area.upper()} TESTS", len(headers))
        current_row += 1
        for tc in cases:
            ws1.cell(row=current_row, column=1, value=f"{prefix}{tc_counter:03d}")
            ws1.cell(row=current_row, column=2, value=tc.scenario)
            ws1.cell(row=current_row, column=3, value=tc.case_type)
            ws1.cell(row=current_row, column=4, value=tc.precondition)
            ws1.cell(row=current_row, column=5, value=tc.steps)
            ws1.cell(row=current_row, column=6, value=tc.expected)
            ws1.cell(row=current_row, column=7, value="")
            ws1.cell(row=current_row, column=8, value="")
            style_body(ws1, current_row, len(headers))
            current_row += 1
            tc_counter += 1

    add_casetype_validation(ws1, f"C2:C{current_row - 1}")
    apply_conditional_fmt(ws1, f"C2:C{current_row - 1}")

    def _kw(*keywords):
        return sum(1 for tc in test_cases if any(kw.lower() in tc.scenario.lower() for kw in keywords))

    summary_rows = [
        ("TOTAL TEST CASES", len(test_cases)),
        ("", ""),
        ("  Positive",  sum(1 for tc in test_cases if tc.case_type == "Positive")),
        ("  Negative",  sum(1 for tc in test_cases if tc.case_type == "Negative")),
        ("  Boundary",  sum(1 for tc in test_cases if tc.case_type == "Boundary")),
        ("", ""),
        ("FORM/UI COMPONENT VALIDATION", ""),
        *[(f"  {area} Component Tests", len(cases)) for area, cases in grouped.items()],
        ("", ""),
        ("NON-FUNCTIONAL TESTING (ESTIMATED)", ""),
        ("  Accessibility (WCAG 2.1 AA)",  _kw("Accessibility", "WCAG", "Contrast")),
        ("  Cross-Browser Compatibility",  _kw("Browser", "Chrome", "Safari")),
        ("  Mobile Responsive Layout",     _kw("Mobile", "Responsive", "Viewport")),
        ("  Performance & Latency",        _kw("Performance", "Slow 3G", "Timeout")),
        ("  Security & Sanitization",      _kw("Security", "XSS", "Injection")),
        ("", ""),
        ("TEST PLAN LANGUAGE", "English"),
        ("GENERATION ENGINE", f"{model_name} (Himagent AI)"),
    ]
    build_summary_sheet(wb, sheet1_title, summary_rows)

    return current_row - 2, ordered_cases


# ── Conditional Formatting ────────────────────────────────────────────────────

def apply_conditional_fmt(ws_testcases, case_col_range: str):
    """
    Apply color-coded conditional formatting:
      - Case Type column: Positive=green, Negative=red, Boundary=yellow
    """
    ws_testcases.conditional_formatting.add(
        case_col_range, CellIsRule(operator='equal', formula=['"Positive"'], fill=green_fill))
    ws_testcases.conditional_formatting.add(
        case_col_range, CellIsRule(operator='equal', formula=['"Negative"'], fill=red_fill))
    ws_testcases.conditional_formatting.add(
        case_col_range, CellIsRule(operator='equal', formula=['"Boundary"'], fill=yellow_fill))
