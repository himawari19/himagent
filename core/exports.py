import csv
import html
import json
from io import BytesIO, StringIO

import openpyxl


def _ws_to_csv_bytes(ws):
    si = StringIO()
    cw = csv.writer(si)
    for row in ws.iter_rows(values_only=True):
        if any(x is not None for x in row):
            cw.writerow([x if x is not None else "" for x in row])
    return si.getvalue().encode('utf-8')


def _ws_to_json_bytes(ws):
    test_cases = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        cells = [x if x is not None else "" for x in row]
        if not any(cells):
            continue
        if not headers:
            if "TC-ID" in cells or "NO" in cells:
                headers = [str(c).upper().replace("-", "_").replace(" ", "_") for c in cells]
            continue
        tc_id = str(cells[0]).strip()
        if tc_id.startswith("TC"):
            tc_dict = {}
            for idx, val in enumerate(cells):
                if idx < len(headers) and headers[idx]:
                    tc_dict[headers[idx].lower()] = val
            test_cases.append(tc_dict)
    return json.dumps(test_cases, indent=2).encode('utf-8')


def _ws_to_markdown_bytes(ws, base_name):
    lines = [f"# Test Plan - {base_name.replace('testplan_', '').replace('_', ' ').title()}", ""]
    headers = []
    col_count = 0
    in_table = False
    for row in ws.iter_rows(values_only=True):
        cells = [str(x) if x is not None else "" for x in row]
        if not any(cells):
            continue
        is_section = (
            cells[0].startswith("SECTION")
            or (len(cells) > 1 and cells[0] == "" and cells[1].startswith("SECTION"))
            or (len(cells) > 1 and "SECTION" in str(cells[0]).upper())
        )
        if is_section:
            sec_title = cells[0] if cells[0] else cells[1]
            if in_table:
                in_table = False
                lines.append("")
            lines.extend([f"## {sec_title}", ""])
            continue
        if "TC-ID" in cells or "NO" in cells:
            headers = [c for c in cells if c]
            col_count = len(headers)
            if in_table:
                lines.append("")
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("| " + " | ".join(["---"] * col_count) + " |")
            in_table = True
            continue
        if in_table and cells[0].strip().startswith("TC"):
            row_cells = cells[:col_count]
            row_cells = [c.replace("|", "\\|").replace("\n", "<br>") for c in row_cells]
            if len(row_cells) < col_count:
                row_cells.extend([""] * (col_count - len(row_cells)))
            lines.append("| " + " | ".join(row_cells) + " |")
    return "\n".join(lines).encode('utf-8')


def _ws_to_html_bytes(ws, base_name):
    title_text = base_name.replace('testplan_', '').replace('_', ' ').title()
    safe_title = html.escape(title_text)
    parts = [
        "<!DOCTYPE html>", "<html>", "<head>", "<meta charset='utf-8'>",
        f"<title>{safe_title} - Himagent Export</title>", "<style>",
        """
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 24px; background: #0f172a; color: #cbd5e1; }
        h1 { color: #f8fafc; border-bottom: 2px solid #334155; padding-bottom: 8px; }
        h2 { color: #38bdf8; margin-top: 32px; border-left: 4px solid #38bdf8; padding-left: 8px; }
        table { width: 100%; border-collapse: collapse; margin-top: 16px; margin-bottom: 24px; background: #1e293b; border-radius: 8px; overflow: hidden; }
        th { background: #2563eb; color: #ffffff; text-align: left; padding: 12px; font-weight: 600; font-size: 14px; }
        td { padding: 12px; border-bottom: 1px solid #334155; font-size: 13px; vertical-align: top; white-space: pre-line; }
        tr:hover { background: #334155; }
        .badge { padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 11px; display: inline-block; }
        .badge-positive { background: #15803d; color: #bbf7d0; }
        .badge-negative { background: #b91c1c; color: #fecaca; }
        .badge-boundary { background: #a16207; color: #fef08a; }
        """,
        "</style>", "</head>", "<body>", f"<h1>Test Plan: {safe_title}</h1>",
    ]
    headers = []
    col_count = 0
    in_table = False
    for row in ws.iter_rows(values_only=True):
        cells = [str(x) if x is not None else "" for x in row]
        if not any(cells):
            continue
        is_section = (
            cells[0].startswith("SECTION")
            or (len(cells) > 1 and cells[0] == "" and cells[1].startswith("SECTION"))
            or (len(cells) > 1 and "SECTION" in str(cells[0]).upper())
        )
        if is_section:
            sec_title = cells[0] if cells[0] else cells[1]
            if in_table:
                parts.append("</table>")
                in_table = False
            parts.append(f"<h2>{html.escape(sec_title)}</h2>")
            continue
        if "TC-ID" in cells or "NO" in cells:
            headers = [c for c in cells if c]
            col_count = len(headers)
            if in_table:
                parts.append("</table>")
            parts.extend(["<table>", "<tr>"])
            for h in headers:
                parts.append(f"<th>{html.escape(h)}</th>")
            parts.append("</tr>")
            in_table = True
            continue
        if in_table and cells[0].strip().startswith("TC"):
            parts.append("<tr>")
            for idx in range(col_count):
                val = cells[idx] if idx < len(cells) else ""
                if idx == 2 and val in ["Positive", "Negative", "Boundary"]:
                    parts.append(f"<td><span class='badge badge-{val.lower()}'>{html.escape(val)}</span></td>")
                else:
                    parts.append(f"<td>{html.escape(val)}</td>")
            parts.append("</tr>")
    if in_table:
        parts.append("</table>")
    parts.extend(["</body>", "</html>"])
    return "\n".join(parts).encode('utf-8')


def export_workbook(file_path, filename, export_format):
    """Convert an xlsx file to the requested format.

    Returns (data, mimetype, download_name) where data is bytes or a file path
    (str) for the xlsx pass-through case.
    """
    base_name = filename.rsplit('.', 1)[0]

    if export_format == 'xlsx':
        return file_path, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    if export_format == 'csv':
        return BytesIO(_ws_to_csv_bytes(ws)), 'text/csv', f"{base_name}.csv"

    if export_format == 'json':
        return BytesIO(_ws_to_json_bytes(ws)), 'application/json', f"{base_name}.json"

    if export_format == 'markdown':
        return BytesIO(_ws_to_markdown_bytes(ws, base_name)), 'text/markdown', f"{base_name}.md"

    if export_format == 'html':
        return BytesIO(_ws_to_html_bytes(ws, base_name)), 'text/html', f"{base_name}.html"

    return None, None, None
