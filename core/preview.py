import os
import threading

import openpyxl

from core.jobs import _GENERATION_LOCK

_PREVIEW_CACHE = {}


def _preview_cache_key(file_path):
    stat = os.stat(file_path)
    return f"{file_path}|{stat.st_mtime_ns}|{stat.st_size}"


def _load_cached(key):
    with _GENERATION_LOCK:
        return _PREVIEW_CACHE.get(key)


def _store_cached(key, result):
    with _GENERATION_LOCK:
        _PREVIEW_CACHE[key] = result
    return result


def _parse_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        max_r = min(ws.max_row or 1, 160)
        max_c = min(ws.max_column or 1, 18)
        for r in range(1, max_r + 1):
            row_cells = []
            row_has_val = False
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                val_str = str(val) if val is not None else ""
                if val_str:
                    row_has_val = True

                bg_color = None
                if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                    rgb = cell.fill.start_color.rgb
                    if isinstance(rgb, str) and rgb not in ('00000000', '000000'):
                        bg_color = "#" + (rgb[2:] if len(rgb) == 8 else rgb)

                is_bold = False
                font_color = None
                if cell.font:
                    is_bold = bool(cell.font.bold)
                    if cell.font.color and cell.font.color.rgb:
                        f_rgb = cell.font.color.rgb
                        if isinstance(f_rgb, str) and f_rgb not in ('00000000', '000000'):
                            font_color = "#" + (f_rgb[2:] if len(f_rgb) == 8 else f_rgb)

                row_cells.append({
                    'value': val_str,
                    'bg_color': bg_color,
                    'font_color': font_color,
                    'is_bold': is_bold,
                    'coordinate': cell.coordinate,
                })
            if row_has_val:
                rows.append(row_cells)
        sheets[sheet_name] = {'rows': rows, 'max_row': len(rows), 'max_col': max_c}
    wb.close()
    return {'success': True, 'type': 'excel', 'sheets': sheets}


def get_file_preview(file_path, filename):
    """Parse a .py or .xlsx file and return a preview dict (cached by mtime)."""
    key = _preview_cache_key(file_path)
    cached = _load_cached(key)
    if cached:
        return cached

    if filename.endswith('.py'):
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        result = {'success': True, 'type': 'python', 'content': content}
        return _store_cached(key, result)

    if filename.endswith('.xlsx'):
        result = _parse_xlsx(file_path)
        return _store_cached(key, result)

    return {'success': False, 'message': 'Unsupported file format.'}
